import os
import argparse
import pandas as pd
from main import read_input
from agents import ai_processor

def format_excel(file_path):
    """把生成的格子调大一点并且加粗表头"""
    import openpyxl
    from openpyxl.styles import Font
    
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.column_dimensions['A'].width = 30  # 公司名称
        ws.column_dimensions['B'].width = 25  # 高管姓名
        ws.column_dimensions['C'].width = 35  # 高管职务
        ws.column_dimensions['D'].width = 40  # 邮件联系方式
        
        # 加粗第一行表头
        bold_font = Font(bold=True)
        for cell in ws[1]:
            if cell.value:
                cell.font = bold_font
                
    wb.save(file_path)

def main():
    parser = argparse.ArgumentParser(description="FirLogic Sales AI - Direct Information Extraction")
    default_output = os.path.expanduser("~/Downloads/FirLogic_Direct_Extraction_Report.xlsx")
    parser.add_argument('--input', type=str, required=True, help="Input file path (txt, docx, or csv/xlsx)")
    parser.add_argument('--output', type=str, default=default_output, help=f"Output file path (default: {default_output})")
    args = parser.parse_args()

    input_path = os.path.expanduser(args.input)
    if not os.path.exists(input_path):
         print(f"Error: 找不到 {input_path}。请检查路径。")
         return
         
    print(f"\n[*] 正在读取输入文件: {input_path}")
    try:
        raw_text = read_input(input_path)
    except Exception as e:
        print(f"[!] 读取文本失败: {e}")
        return

    print(f"    成功读取内容长度: {len(raw_text)} 字符。")
    print("\n---------------------------------------------------------")
    print("[*] 启动跳板车大模型 (Direct LLM Extraction)")
    print("---------------------------------------------------------")
    
    results = ai_processor.run_direct_extraction(raw_text)
    
    if not results:
        print("\n[!] 警告: 没有提取到任何有效人员信息。")
        return
        
    print(f"\n[*] 成功提取 {len(results)} 位高管联系人。正在写入报表...")
    
    # 构建 DataFrame，调整列的顺序和中文表头以对接后续步骤或作为最终成品
    formatted_results = []
    for p in results:
        formatted_results.append({
            '公司名称': p.get('company_name', ''),
            '高管姓名': p.get('name', ''),
            '高管职务': p.get('title', ''),
            '邮件联系方式': p.get('email', '')
        })
        
    df = pd.DataFrame(formatted_results)
    
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='查询结果', index=False)
        
    print("[*] 正在精美重排(拉宽单元格)... ", end="", flush=True)
    try:
        format_excel(args.output)
        print("完成")
    except Exception as e:
        print(f"失败 ({e})")
        
    print(f"\n[OK] 一步提取完成！最终提取报表已保存至: {args.output}")

if __name__ == "__main__":
    main()
