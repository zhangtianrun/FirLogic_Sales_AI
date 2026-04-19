import os
import argparse
import pandas as pd
from agents import ai_processor

def build_email_template(company, salutation, last_name, location):
    subject = f"A new log scaling technology from Ontario Canada — Up to 80% cost reduction for {company}"
    body = f"""Dear {salutation} {last_name},

{company}’s long-standing commitment to sustainable forestry and operational excellence is truly impressive.

My name is Terry Zhang, and I am the General Manager of Fir Logic, a Canadian technology company established at the University of Waterloo and stationed in Ontario. I am reaching out today because I believe we can greatly support you by modernizing your log scaling process and removing one of the industry's biggest bottlenecks.

Fir Logic specializes in a patented 3D vision-based portable AI scaling solution that removes the reliance on manual scaling, delivers objective results, and reduces traditional operational costs by up to 80%. For more details, please review the attached document.

For more information on our projects and transparent pricing, please visit our website: https://www.firlogic.com/pricing 

If you have any questions or would like to learn more, please don't hesitate to contact us. If you are interested, we would be delighted to collaborate with you.

Best regards,
Terry Zhang 
General Manager, Fir Logic Ltd."""
    return subject, body

def format_excel(file_path):
    """设置特定的列宽并严禁换行"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.column_dimensions['A'].width = 25  # 公司名称
        ws.column_dimensions['B'].width = 20  # 高管姓名
        ws.column_dimensions['C'].width = 30  # 邮件联系方式
        ws.column_dimensions['D'].width = 20  # 公司所在地
        ws.column_dimensions['E'].width = 50  # Subject
        ws.column_dimensions['F'].width = 60  # Email Body
        
        bold_font = Font(bold=True)
        # 单行不换行，且靠上对齐
        no_wrap_alignment = Alignment(wrap_text=False, vertical='top')
        
        # 加粗第一行
        for cell in ws[1]:
            if cell.value:
                cell.font = bold_font
                cell.alignment = no_wrap_alignment
                
        # 设置所有数据格的对齐方式，防止邮件正文把格子撑大
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = no_wrap_alignment
                
    wb.save(file_path)

def main():
    parser = argparse.ArgumentParser(description="FirLogic Sales AI - Step 4: Email Drafter")
    default_output = os.path.expanduser("~/Downloads/FirLogic_Sales_Intel_Report_Step4.xlsx")
    parser.add_argument('--input', type=str, default="~/Downloads/FirLogic_Sales_Intel_Report_Step3.xlsx", help="Step 3 输出文件路径")
    parser.add_argument('--output', type=str, default=default_output, help=f"生成 Step 4 的输出文件路径")
    args = parser.parse_args()

    input_file = os.path.expanduser(args.input)
    output_file = os.path.expanduser(args.output)

    if not os.path.exists(input_file):
        print(f"Error: 找不到 {input_file}。请检查路径。")
        return

    print("\n---------------------------------------------------------")
    print("[*] 启动写信模式 (Email Drafter Pipeline)")
    print("---------------------------------------------------------")
    print(f"[*] 正在读取输入文件: {input_file}")

    try:
        # 支持从任何 Sheet 中读取，默认优先读取 '查询结果'
        xl = pd.ExcelFile(input_file)
        sheet_name = '查询结果' if '查询结果' in xl.sheet_names else xl.sheet_names[0]
        df = pd.read_excel(input_file, sheet_name=sheet_name)
    except Exception as e:
        print(f"\n[!] 读取 Excel 失败: {e}")
        return

    results = []
    skipped_count = 0
    
    # 公司地理位置缓存：一家公司只查一次 Google，省钱、提速、保一致性
    company_location_cache = {}

    for idx, row in df.iterrows():
        company = str(row.get('公司名称', '')).strip()
        name = str(row.get('高管姓名', '')).strip()
        email = str(row.get('邮件联系方式', '')).strip()

        # 核心过滤逻辑：只给已经拿到真实邮箱的人写信
        invalid_markers = ["no result", "跳过", "拦截", "超时", "失败", "nan"]
        is_invalid = any(marker in email.lower() for marker in invalid_markers) or "@" not in email
        
        if is_invalid or not company or not name or not email:
            skipped_count += 1
            continue
            
        # 确定域名锚点：优先从表格读，没有则从邮箱反推
        raw_website = str(row.get('公司网站', row.get('官方网站', row.get('Website', '')))).strip()
        domain = raw_website
        
        if not domain or domain.lower() == 'nan':
            # 从邮箱反推域名 (例如: dave@inglewoodsawmill.com.au -> inglewoodsawmill.com.au)
            if "@" in email:
                domain = email.split("@")[-1].strip()
        
        # 1. 获取地理位置（查缓存或查 AI）
        if company not in company_location_cache:
            location = ai_processor.run_company_location_research(company, domain)
            company_location_cache[company] = location
        else:
            location = company_location_cache[company]
            
        # 2. 获取个人身份信息（不联网 AI）
        identity = ai_processor.run_identity_analysis(name)
        
        salutation = identity.get('salutation', 'Mr.')
        last_name = identity.get('last_name', name.split()[-1] if name.split() else name)
        
        # 组装邮件
        subject, body = build_email_template(company, salutation, last_name, location)
        
        results.append({
            '公司名称': company,
            '高管姓名': name,
            '邮件联系方式': email,
            '公司所在地': location,
            'Subject': subject,
            'Email body': body
        })

    print(f"\n[*] 生成完毕。有效生成: {len(results)} 封，跳过无效/无邮箱数据: {skipped_count} 条。")
    
    if not results:
        print("[-] 警告: 您提供的名单里没有有效的可群发邮件对象。")
        return

    print(f"[*] 正在写入报表 {output_file} ...", end="", flush=True)
    os.makedirs(os.path.dirname(os.path.abspath(output_file)), exist_ok=True)
    
    df_out = pd.DataFrame(results)
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name='开发信库', index=False)
    print(" 完成")

    print("[*] 正在执行防变形排版 (严禁换行)...", end="", flush=True)
    try:
        format_excel(output_file)
        print(" 完成")
    except Exception as e:
        print(f" 失败 ({e})")
        
    print(f"\n[OK] 终极冷邮件弹药库已准备完毕！请前往 {output_file} 提取。")

if __name__ == "__main__":
    main()
