import argparse
import os
import pandas as pd
from docx import Document
from core.pipeline import process_leads

def read_input(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Input file {file_path} not found.")
        
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ['.txt', '.csv']:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    elif ext == '.docx':
        doc = Document(file_path)
        return "\n".join([para.text for para in doc.paragraphs])
    elif ext in ['.xlsx', '.xls']:
        df = pd.read_excel(file_path)
        # Convert the whole dataframe to text to pass to the extractor
        return df.to_string()
    else:
        raise ValueError("Unsupported valid format. Use txt, docx, csv, or xlsx.")

def flatten_results(results_list):
    flat_list = []
    base_keys = ["公司名称", "网站", "木材类别", "人员数量", "厂数量", "业务分类", "具体品种", "自动化程度", "竞品设备", "理由"]
    for item in results_list:
        staff_list = item.pop("staff_list", [])
        if not staff_list:
            flat_item = {k: item.get(k, "") for k in base_keys}
            flat_item["高管姓名"] = "未找到"
            flat_item["高管职务"] = "未找到"
            flat_item["职责描述"] = "未找到"
            flat_item["销售分析"] = "未找到"
            flat_item["情报来源链接"] = "未找到"
            flat_list.append(flat_item)
        else:
            for i, staff in enumerate(staff_list):
                flat_item = {}
                for k in base_keys:
                    if i == 0:
                        flat_item[k] = item.get(k, "")
                    else:
                        flat_item[k] = "" # Leave empty for merging
                flat_item["高管姓名"] = staff.get("name", "")
                flat_item["高管职务"] = staff.get("title", "")
                flat_item["职责描述"] = staff.get("role_description", "")
                flat_item["销售分析"] = staff.get("relevance_analysis", "")
                flat_item["情报来源链接"] = staff.get("source_link", "")
                flat_list.append(flat_item)
    return flat_list

def write_to_excel(results, output_path):
    from openpyxl.styles import Font, Alignment
    
    target_softwood = []
    target_hardwood = []
    excluded = []
    
    for r in results:
        # Copy to avoid modifying the original objects if they are reused
        item = r.copy()
        tab = item.pop("__tab__", "Excluded")
        wood_raw = item.pop("__wood_raw__", "").lower()
        
        if tab == "Target":
            is_mixed = "混合" in wood_raw or "mixed" in wood_raw
            is_hardwood = "硬木" in wood_raw or "hardwood" in wood_raw
            
            if is_mixed:
                target_softwood.append(item)
            elif is_hardwood:
                target_hardwood.append(item)
            else:
                target_softwood.append(item)
        else:
            excluded.append(item)
            
    df_softwood = pd.DataFrame(flatten_results(target_softwood)) if target_softwood else pd.DataFrame()
    df_hardwood = pd.DataFrame(flatten_results(target_hardwood)) if target_hardwood else pd.DataFrame()
    df_excluded = pd.DataFrame(excluded) if excluded else pd.DataFrame()
    
    # Updated Column Order (Chinese)
    cols_target = [
        "公司名称", "网站", "木材类别", "人员数量", "厂数量", 
        "高管姓名", "高管职务", "职责描述", "销售分析", "情报来源链接",
        "业务分类", "具体品种", "自动化程度", "竞品设备", "理由"
    ]
    
    cols_excluded = ["公司名称", "业务分类", "理由", "网站"]
    
    # Reorder columns
    if not df_softwood.empty: df_softwood = df_softwood.reindex(columns=cols_target)
    if not df_hardwood.empty: df_hardwood = df_hardwood.reindex(columns=cols_target)
    if not df_excluded.empty: df_excluded = df_excluded.reindex(columns=cols_excluded)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_softwood.to_excel(writer, sheet_name="重点关注_软木及混合", index=False)
        df_hardwood.to_excel(writer, sheet_name="重点关注_硬木", index=False)
        df_excluded.to_excel(writer, sheet_name="非目标_已剔除", index=False)
        
        workbook = writer.book
        bold_font = Font(bold=True)
        
        for sheetname in writer.sheets:
            worksheet = writer.sheets[sheetname]
            # 1. Bold the header row
            for cell in worksheet[1]:
                cell.font = bold_font
            
            # 2. Adjust column widths and text wraps
            header_map = {cell.value: cell.column_letter for cell in worksheet[1] if cell.value}
            
            # Default widths for all columns
            for col in worksheet.columns:
                if col[0].value:
                    col_letter = col[0].column_letter
                    worksheet.column_dimensions[col_letter].width = 25
            
            # Specific wider columns
            for hdr in ["理由", "自动化程度", "竞品设备", "具体品种", "职责描述", "销售分析"]:
                if hdr in header_map:
                    worksheet.column_dimensions[header_map[hdr]].width = 45
            for hdr in ["情报来源链接"]:
                if hdr in header_map:
                    worksheet.column_dimensions[header_map[hdr]].width = 15

            # Apply Merge Cells for base columns if they are empty
            merge_cols = ["公司名称", "网站", "木材类别", "人员数量", "厂数量", "业务分类", "具体品种", "自动化程度", "竞品设备", "理由"]
            for col_name in merge_cols:
                if col_name in header_map:
                    col_letter = header_map[col_name]
                    col_idx = worksheet[col_letter + "1"].column
                    
                    start_row = 2
                    while start_row <= worksheet.max_row:
                        cell_val = worksheet.cell(row=start_row, column=col_idx).value
                        if cell_val != "" and cell_val is not None:
                            end_row = start_row
                            while end_row + 1 <= worksheet.max_row and (worksheet.cell(row=end_row + 1, column=col_idx).value == "" or worksheet.cell(row=end_row + 1, column=col_idx).value is None):
                                end_row += 1
                            if end_row > start_row:
                                worksheet.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                                worksheet.cell(row=start_row, column=col_idx).alignment = Alignment(vertical='center', wrap_text=True)
                            else:
                                worksheet.cell(row=start_row, column=col_idx).alignment = Alignment(vertical='top', wrap_text=True)
                            start_row = end_row + 1
                        else:
                            start_row += 1
                            
            # Enable word wrap for these columns explicitly
            for col_name in ["职责描述", "销售分析", "高管姓名", "高管职务"]:
                if col_name in header_map:
                    col_letter = header_map[col_name]
                    for row in range(2, worksheet.max_row + 1):
                        worksheet[f"{col_letter}{row}"].alignment = Alignment(vertical='top', wrap_text=True)
            
            # 3. Bold the wood category column content
            if "木材类别" in header_map:
                col_letter = header_map["木材类别"]
                for row in range(2, worksheet.max_row + 1):
                    worksheet[f"{col_letter}{row}"].font = bold_font

    print(f"\n[Success] Report generated: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="Fir Logic - AI Sales Intel Generator (AI Grounding Edition)")
    default_output = os.path.expanduser("~/Downloads/FirLogic_Sales_Intel_Report.xlsx")
    parser.add_argument('--input', type=str, required=True, help="Input file path (txt, docx, or xlsx)")
    parser.add_argument('--output', type=str, default=default_output, help=f"Output file path (default: {default_output})")
    args = parser.parse_args()
    
    print(f"Reading input from {args.input}...")
    try:
        raw_text = read_input(args.input)
    except Exception as e:
        print(f"Error reading file: {e}")
        return
        
    results = process_leads(raw_text)
    
    print(f"\nWriting {len(results)} records to Excel...")
    write_to_excel(results, args.output)

if __name__ == "__main__":
    main()
