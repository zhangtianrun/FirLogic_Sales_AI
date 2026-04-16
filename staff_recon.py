import argparse
import os
import pandas as pd
from agents import ai_processor
import config

def flatten_results(results_list):
    flat_list = []
    base_keys = ["公司名称", "网站", "木材类别", "人员数量", "厂数量", "业务分类", "具体品种", "自动化程度", "竞品设备", "理由"]
    for item in results_list:
        staff_list = item.pop("staff_list", [])
        if not staff_list:
            flat_item = {k: item.get(k, "") for k in base_keys}
            flat_item["高管姓名"] = "未找到"
            flat_item["高管职务"] = "未找到"
            flat_item["职责描述"] = "未找到"
            flat_item["销售分析"] = "未找到"
            flat_item["情报来源链接"] = "未找到"
            flat_list.append(flat_item)
        else:
            for i, staff in enumerate(staff_list):
                flat_item = {}
                for k in base_keys:
                    if i == 0:
                        flat_item[k] = item.get(k, "")
                    else:
                        flat_item[k] = "" # Leave empty for merging
                flat_item["高管姓名"] = staff.get("name", "")
                flat_item["高管职务"] = staff.get("title", "")
                flat_item["职责描述"] = staff.get("role_description", "")
                flat_item["销售分析"] = staff.get("relevance_analysis", "")
                flat_item["情报来源链接"] = staff.get("source_link", "")
                flat_list.append(flat_item)
    return flat_list

def process_and_export(input_path, output_path):
    print(f"Reading targets from {input_path}...")
    
    # Read the two target sheets if they exist
    target_results = []
    
    try:
        xl = pd.ExcelFile(input_path)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return
        
    for sheet_name in ["重点关注_软木及混合", "重点关注_硬木"]:
        if sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name)
            # convert NaNs to empty strings
            df = df.fillna("")
            
            for index, row in df.iterrows():
                company_name = row.get("公司名称", "")
                if not company_name:
                    continue
                    
                # Convert row to dictionary
                info = row.to_dict()
                info["__tab__"] = sheet_name
                
                print(f"\n[Staff Recon] Penetrating {company_name}...")
                staff = ai_processor.run_staff_test(company_name, config.MODEL_DETECTIVE)
                info["staff_list"] = staff
                target_results.append(info)

    if not target_results:
        print("No targets found in the input Excel. Did you process the right file?")
        return
        
    print("\nGenerating final formatting Excel...")
    
    target_softwood = [r for r in target_results if r.get("__tab__") == "重点关注_软木及混合"]
    target_hardwood = [r for r in target_results if r.get("__tab__") == "重点关注_硬木"]
    
    df_softwood = pd.DataFrame(flatten_results(target_softwood)) if target_softwood else pd.DataFrame()
    df_hardwood = pd.DataFrame(flatten_results(target_hardwood)) if target_hardwood else pd.DataFrame()
    
    # Updated Column Order (Chinese)
    cols_target = [
        "公司名称", "网站", "木材类别", "人员数量", "厂数量", 
        "高管姓名", "高管职务", "职责描述", "销售分析", "情报来源链接",
        "业务分类", "具体品种", "自动化程度", "竞品设备", "理由"
    ]
    
    if not df_softwood.empty: df_softwood = df_softwood.reindex(columns=cols_target)
    if not df_hardwood.empty: df_hardwood = df_hardwood.reindex(columns=cols_target)
    
    from openpyxl.styles import Font, Alignment
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if not df_softwood.empty or not df_hardwood.empty:
            if not df_softwood.empty:
                df_softwood.to_excel(writer, sheet_name="重点关注_软木及混合", index=False)
            if not df_hardwood.empty:
                df_hardwood.to_excel(writer, sheet_name="重点关注_硬木", index=False)
            
            workbook = writer.book
            bold_font = Font(bold=True)
            
            for sheetname in writer.sheets:
                worksheet = writer.sheets[sheetname]
                # 1. Bold the header row
                for cell in worksheet[1]:
                    if cell.value:
                        cell.font = bold_font
                
                # 2. Adjust column widths and text wraps
                header_map = {cell.value: cell.column_letter for cell in worksheet[1] if cell.value}
                
                # Default widths for all columns
                for col in worksheet.columns:
                    if col[0].value:
                        col_letter = col[0].column_letter
                        worksheet.column_dimensions[col_letter].width = 25
                
                # Specific wider columns
                for hdr in ["理由", "自动化程度", "竞品设备", "具体品种", "职责描述", "销售分析"]:
                    if hdr in header_map:
                        worksheet.column_dimensions[header_map[hdr]].width = 45
                for hdr in ["情报来源链接"]:
                    if hdr in header_map:
                        worksheet.column_dimensions[header_map[hdr]].width = 15

                # Apply Merge Cells for base columns if they are empty
                merge_cols = ["公司名称", "网站", "木材类别", "人员数量", "厂数量", "业务分类", "具体品种", "自动化程度", "竞品设备", "理由"]
                for col_name in merge_cols:
                    if col_name in header_map:
                        col_letter = header_map[col_name]
                        col_idx = worksheet[col_letter + "1"].column
                        
                        start_row = 2
                        while start_row <= worksheet.max_row:
                            cell_val = worksheet.cell(row=start_row, column=col_idx).value
                            if cell_val != "" and cell_val is not None:
                                end_row = start_row
                                while end_row + 1 <= worksheet.max_row and (worksheet.cell(row=end_row + 1, column=col_idx).value == "" or worksheet.cell(row=end_row + 1, column=col_idx).value is None):
                                    end_row += 1
                                if end_row > start_row:
                                    worksheet.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                                    worksheet.cell(row=start_row, column=col_idx).alignment = Alignment(vertical='center', wrap_text=True)
                                else:
                                    worksheet.cell(row=start_row, column=col_idx).alignment = Alignment(vertical='top', wrap_text=True)
                                start_row = end_row + 1
                            else:
                                start_row += 1
                                
                # Enable word wrap for these columns explicitly
                for col_name in ["职责描述", "销售分析", "高管姓名", "高管职务"]:
                    if col_name in header_map:
                        col_letter = header_map[col_name]
                        for row in range(2, worksheet.max_row + 1):
                            worksheet[f"{col_letter}{row}"].alignment = Alignment(vertical='top', wrap_text=True)
                
                # 3. Bold the wood category column content
                if "木材类别" in header_map:
                    col_letter = header_map["木材类别"]
                    for row in range(2, worksheet.max_row + 1):
                        worksheet[f"{col_letter}{row}"].font = bold_font

    print(f"\n[Success] Final Report generated with Staff Intel: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="Fir Logic - AI Sales Intel Generator (Staff Recon Edition)")
    default_input = os.path.expanduser("~/Downloads/FirLogic_Sales_Intel_Report_Step1.xlsx")
    default_output = os.path.expanduser("~/Downloads/FirLogic_Sales_Intel_Report_Step2.xlsx")
    
    parser.add_argument('--input', type=str, default=default_input, help=f"Input file path (default: {default_input})")
    parser.add_argument('--output', type=str, default=default_output, help=f"Output file path (default: {default_output})")
    args = parser.parse_args()
    
    process_and_export(args.input, args.output)

if __name__ == "__main__":
    main()
